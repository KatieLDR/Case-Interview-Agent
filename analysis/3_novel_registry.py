#!/usr/bin/env python3
"""3_novel_registry.py — scored workbook -> blind two-rater creativity registry.

Step 3 of the analysis-scoring pipeline (see ba-agent/docs/analysis_scoring_plan.md).

Runs AFTER the human finishes Round 2 tagging in scoring.xlsx. Collects every item
tagged R2_tag == "novel", dedupes exact-normalized duplicates into one novel_id, and
emits, into analysis/data/:
  - novel_ideas_raters_<collection>_<date>.xlsx  the BLIND file two raters fill
                                       sequentially (rate1/comment1 -> same file to
                                       rater 2 -> rate2/comment2), with a live
                                       rate1&rate2 agreement column and a manual
                                       consensus final_decision. Raters see ONLY the
                                       idea text — no participant/session/arm columns.
  - novel_crosswalk_<collection>_<date>.csv  novel_id -> (session_id, participant_id,
                                       item_index, idea_text). The researcher-only
                                       merge-back key store (the analogue of script 2's
                                       crosswalk.csv); used by 4_merge.py only.

Design guarantees (mirror the plan's acceptance checklist, section 3):
  * READ-ONLY on the scored workbook — opened, never modified. No Firestore, no backend/.
  * NO LLM. Dedup is exact-normalized string matching; creativity is human judgement.
  * BLIND. The raters' xlsx carries NO participant/session/arm column. Identity lives
    only in novel_crosswalk.csv.
  * NON-DESTRUCTIVE. Refuses to overwrite an existing raters xlsx (a returned file may
    hold rater input) or crosswalk. Re-run with a fresh date stamp or remove the file.

Method caveat (record in the thesis): the two raters work sequentially on ONE shared
file, so rater 2 is not blind to rater 1's ratings. Inter-rater agreement (reported by
4_merge.py) must be described as non-independent.
"""

import argparse
import csv
import datetime as _dt
import difflib
import glob
import os
import re
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(HERE)  # ba-agent/
DATA_DIR = os.path.join(HERE, "data")

# The R2 tag we harvest. Must match the literal the Composite tab counts in script 2
# (Items!$H:$H,"novel").
NOVEL_TAG = "novel"

# Columns script 2 writes into the Items sheet; we read a subset by name (never by
# letter) so a future column reorder in scoring.xlsx can't silently mis-map values.
REQUIRED_ITEMS_COLS = ["session_id", "participant_id", "item_index", "item_text",
                       "apparatus_swap", "R2_tag"]


# ── normalisation / dedup ────────────────────────────────────────────────────
def _norm(text: str) -> str:
    """Exact-normalized dedup key: lowercased, whitespace-collapsed, edge punct stripped."""
    s = re.sub(r"\s+", " ", (text or "").strip().lower())
    return s.strip(" .;:")


# ── reading the scored workbook ──────────────────────────────────────────────
def read_novel_items(scoring_path):
    """Return the novel-tagged, non-swap items from the Items sheet, in sheet order.

    Each element: {session_id, participant_id, item_index, item_text}. Raises via
    sys.exit if the workbook lacks an Items sheet or a required column.
    """
    wb = load_workbook(scoring_path, read_only=True, data_only=False)
    if "Items" not in wb.sheetnames:
        sys.exit(f"ERROR: {scoring_path} has no 'Items' sheet — is this a scoring workbook "
                 f"produced by 2_build_workbook.py?")
    ws = wb["Items"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        sys.exit(f"ERROR: {scoring_path} 'Items' sheet is empty.")
    col = {name: i for i, name in enumerate(header)}
    missing = [c for c in REQUIRED_ITEMS_COLS if c not in col]
    if missing:
        sys.exit(f"ERROR: Items sheet missing column(s) {missing}. Found: {header}")

    items = []
    for r in rows:
        tag = r[col["R2_tag"]]
        if tag is None or str(tag).strip().lower() != NOVEL_TAG:
            continue
        # Defensive: a planted-swap row is quarantined in BOTH rounds and must never
        # reach the raters, even if mis-tagged novel by hand.
        if r[col["apparatus_swap"]]:
            continue
        text = r[col["item_text"]]
        if text is None or not str(text).strip():
            continue
        items.append({
            "session_id": r[col["session_id"]],
            "participant_id": r[col["participant_id"]] or "",
            "item_index": r[col["item_index"]],
            "item_text": str(text).strip(),
        })
    wb.close()
    return items


def build_registry(items):
    """Group novel items into novel_ids by exact-normalized text (first-seen order).

    Returns (ideas, crosswalk_rows):
      ideas         : list of {novel_id, idea_text, n_participants}
      crosswalk_rows: list of {novel_id, session_id, participant_id, item_index, idea_text}
    """
    order = []          # normalized keys in first-seen order
    by_key = {}         # key -> {novel_id, idea_text, members:[item,...]}
    for it in items:
        key = _norm(it["item_text"])
        if key not in by_key:
            order.append(key)
            by_key[key] = {"idea_text": it["item_text"], "members": []}
        by_key[key]["members"].append(it)

    ideas = []
    crosswalk_rows = []
    for i, key in enumerate(order, start=1):
        nid = f"N{i:03d}"
        grp = by_key[key]
        # session_id is the per-participant identity (unique per doc; the g0 pilot has
        # blank participant_ids, so counting pids would wrongly collapse them — script 2
        # keys its Composite tab on session_id for the same reason).
        sessions = {m["session_id"] for m in grp["members"]}
        ideas.append({
            "novel_id": nid,
            "idea_text": grp["idea_text"],
            "n_participants": len(sessions),
        })
        for m in grp["members"]:
            crosswalk_rows.append({
                "novel_id": nid,
                "session_id": m["session_id"],
                "participant_id": m["participant_id"],
                "item_index": m["item_index"],
                "idea_text": m["item_text"],
            })
    return ideas, crosswalk_rows


def borderline_pairs(ideas, threshold):
    """Distinct novel_id pairs whose texts are near-identical (>= threshold) but not
    exact-normalized equal (those are already one id). Listed for manual merge."""
    out = []
    for a in range(len(ideas)):
        for b in range(a + 1, len(ideas)):
            ratio = difflib.SequenceMatcher(
                None, _norm(ideas[a]["idea_text"]), _norm(ideas[b]["idea_text"])).ratio()
            if ratio >= threshold:
                out.append((ideas[a], ideas[b], ratio))
    out.sort(key=lambda t: t[2], reverse=True)
    return out


# ── raters workbook construction ─────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(bold=True)

RATER_HEADERS = ["novel_id", "idea_text", "rate1", "comment1", "rate2", "comment2",
                 "rate1&rate2", "final_decision"]


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def build_raters_workbook(ideas, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ratings"
    ws.append(RATER_HEADERS)

    for i, idea in enumerate(ideas):
        r = i + 2
        # rate1&rate2: blank until both rated, then TRUE only if both TRUE.
        agree = f'=IF(OR(C{r}="",E{r}=""), "", AND(C{r}=TRUE,E{r}=TRUE))'
        ws.append([
            idea["novel_id"], idea["idea_text"],
            "", "",       # rate1, comment1  (rater 1)
            "", "",       # rate2, comment2  (rater 2)
            agree,        # rate1&rate2      (auto)
            "",           # final_decision   (raters, jointly)
        ])

    _style_header(ws, len(RATER_HEADERS))
    for col, w in zip("ABCDEFGH", (10, 90, 8, 34, 8, 34, 12, 14)):
        ws.column_dimensions[col].width = w

    last = len(ideas) + 1
    if last >= 2:
        for letter, title, prompt in (
            ("C", "rate1", "Rater 1: is this idea genuinely creative/novel? TRUE or FALSE."),
            ("E", "rate2", "Rater 2: is this idea genuinely creative/novel? TRUE or FALSE."),
            ("H", "final_decision",
             "Consensus: agreed value where raters agree; resolve disagreements by "
             "discussion and record the outcome. Must be filled for every row."),
        ):
            dv = DataValidation(
                type="list",
                formula1='"TRUE,FALSE"',
                allow_blank=True,
                showErrorMessage=True,
                showInputMessage=True,
            )
            dv.error = "Choose TRUE or FALSE."
            dv.errorTitle = "Not TRUE/FALSE"
            dv.prompt = prompt
            dv.promptTitle = title
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{last}")

        # idea_text wraps; ratings/decision are short.
        for r in range(2, last + 1):
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # Header guidance comments (mirror script 2's Comment(text, "pipeline")).
    ws.cell(row=1, column=2).comment = Comment(
        "BLIND: raters see only this idea text — no participant/session/arm. Workflow: "
        "rater 1 fills rate1/comment1, then the SAME file goes to rater 2 for "
        "rate2/comment2. Sequential (rater 2 not blind to rater 1) — report agreement as "
        "non-independent.", "pipeline")
    ws.cell(row=1, column=7).comment = Comment(
        "Auto formula: blank until both rate1 and rate2 are filled, then TRUE only if "
        "BOTH raters rated TRUE, else FALSE. Do not edit.", "pipeline")
    ws.cell(row=1, column=8).comment = Comment(
        "Consensus decision. Where raters agree, equals the agreed value; where they "
        "disagree, resolve by discussion and record the outcome here. 4_merge.py refuses "
        "to run while any final_decision is blank.", "pipeline")

    wb.save(out_path)


# ── main ─────────────────────────────────────────────────────────────────────
def _latest_scoring():
    cands = sorted(glob.glob(os.path.join(DATA_DIR, "scoring_*.xlsx")))
    return cands[-1] if cands else None


def _collection_from_name(path):
    """Recover <collection> from a scoring_<collection>_<date>.xlsx filename."""
    m = re.match(r"scoring_(?P<coll>.+)_\d{4}-\d{2}-\d{2}\.xlsx$", os.path.basename(path))
    return m.group("coll") if m else "unknown"


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--scoring", default=None,
                    help="Path to the scored scoring_*.xlsx (default: newest in analysis/data).")
    ap.add_argument("--out-dir", default=DATA_DIR,
                    help="Output directory (default: analysis/data).")
    ap.add_argument("--similarity", type=float, default=0.85,
                    help="Borderline near-duplicate ratio threshold, 0..1 (default: 0.85).")
    args = ap.parse_args()

    scoring_path = args.scoring or _latest_scoring()
    if not scoring_path or not os.path.exists(scoring_path):
        sys.exit("ERROR: no scoring workbook found. Run 2_build_workbook.py first, or "
                 "pass --scoring <path>.")

    collection = _collection_from_name(scoring_path)
    today = _dt.date.today().isoformat()
    xlsx_path = os.path.join(args.out_dir, f"novel_ideas_raters_{collection}_{today}.xlsx")
    cross_path = os.path.join(args.out_dir, f"novel_crosswalk_{collection}_{today}.csv")
    for p in (xlsx_path, cross_path):
        if os.path.exists(p):
            sys.exit(f"ERROR: {p} already exists. The raters registry may hold rater "
                     f"input — it is never overwritten. Remove it manually to rebuild, "
                     f"or wait for tomorrow's date stamp.")

    print(f"Scoring workbook : {scoring_path}")
    print(f"Collection       : {collection}")

    items = read_novel_items(scoring_path)
    if not items:
        sys.exit("No R2_tag == 'novel' items found in the Items sheet. Nothing to rate — "
                 "has Round 2 tagging been done in this workbook?")

    ideas, crosswalk_rows = build_registry(items)
    build_raters_workbook(ideas, xlsx_path)

    with open(cross_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["novel_id", "session_id", "participant_id", "item_index", "idea_text"])
        for cr in crosswalk_rows:
            w.writerow([cr["novel_id"], cr["session_id"], cr["participant_id"],
                        cr["item_index"], cr["idea_text"]])

    # ── report ──────────────────────────────────────────────────────────────
    shared = [i for i in ideas if i["n_participants"] > 1]
    participants = {cr["session_id"] for cr in crosswalk_rows}
    borders = borderline_pairs(ideas, args.similarity)

    print("\n" + "=" * 70)
    print("NOVEL REGISTRY")
    print("=" * 70)
    print(f"novel-tagged items        : {len(items)}")
    print(f"unique novel ideas        : {len(ideas)}  (novel_id N001..N{len(ideas):03d})")
    print(f"ideas shared by >1 person : {len(shared)}")
    for i in shared:
        print(f"   {i['novel_id']}: {i['n_participants']} people (by session_id)")
    print(f"participants involved     : {len(participants)}  (distinct session_id)")
    if borders:
        print(f"\nborderline near-dupes (ratio >= {args.similarity}) — MERGE MANUALLY if "
              f"the same idea (edit scoring.xlsx text to match, then re-run):")
        for a, b, ratio in borders:
            print(f"   {ratio:.2f}  {a['novel_id']} <> {b['novel_id']}")
            print(f"        {a['novel_id']}: {a['idea_text']!r}")
            print(f"        {b['novel_id']}: {b['idea_text']!r}")
    else:
        print(f"\nborderline near-dupes     : none (ratio >= {args.similarity})")

    print("\n" + "=" * 70)
    print(f"Wrote raters registry : {xlsx_path}")
    print(f"Wrote novel crosswalk : {cross_path}")
    print("Registry is BLIND (no participant/session/arm). Identity lives only in the "
          "crosswalk (4_merge.py input).")
    print("=" * 70)


if __name__ == "__main__":
    main()
