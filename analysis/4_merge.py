#!/usr/bin/env python3
"""4_merge.py — merge scores + creativity back into the Qualtrics master.

Step 4 (final) of the analysis-scoring pipeline (see
ba-agent/docs/analysis_scoring_plan.md). This is the integrity gate: it joins the
hand-scored workbook, the two-rater creativity registry, and the Qualtrics export,
re-attaches arm identity, and reconciles all three sources against each other.

Inputs (auto-discovered by collection unless overridden):
  --scoring          scoring_<collection>_<date>.xlsx      (script 2; Items tab scored)
  --crosswalk        crosswalk_<collection>_<date>.csv     (script 2; session->agent_type)
  --raters           novel_ideas_raters_<collection>_<date>.xlsx  (script 3, RETURNED)
  --novel-crosswalk  novel_crosswalk_<collection>_<date>.csv      (script 3; novel_id map)
  --snapshot         raw_snapshot_<collection>_<date>.json (script 1; for the N count)
  --qualtrics        the Qualtrics export (CSV or XLSX)     REQUIRED, user-supplied
  --responseid-col   Qualtrics column holding ResponseId    (default "ResponseId")
  --session-col      Qualtrics column(s) holding the pasted Session ID (enables fallback
                     join). Comma-separated when the Session ID is split across per-arm
                     fields, e.g. "Q23,Q27,Q28" — only one is filled per respondent, so
                     the first non-blank value in a row is used.

Outputs, into analysis/data/ (overwritten each run — they are DERIVED, deterministic):
  master_scored_<collection>_<date>.csv
  reconciliation_report_<collection>_<date>.txt

Design guarantees (mirror the plan's acceptance checklist, sections 0 and 4):
  * READ-ONLY. Opens every input, modifies none. No Firestore. No backend.logger.
    No .set()/.update()/.delete(). No LLM/genai imports. Config comes from CLI args.
  * COMPOSITE UNCHANGED. composite = R1 hits + R2 extension + R2 novel, recomputed
    directly from the Items tab, enforcing R1/R2 as disjoint (an R1 hit is never also
    counted as R2). NOTE: this diverges from the Excel Composite tab's raw COUNTIFS
    for any row where a rater filled BOTH R1_kb_id and R2_tag — Excel's independent
    COUNTIFS would double-count that row; script 4 counts it once (R1) and flags it
    in the reconciliation report so it can be fixed at the source. Creativity does
    NOT gate the composite — it is reported standalone (n_novel_ideas,
    n_creative_confirmed).
  * BLIND UNTIL NOW. agent_type is rejoined from crosswalk.csv only here, after all
    scoring is done.
  * DETERMINISTIC + NON-DESTRUCTIVE. Recomputes from raw item rows (never trusts cached
    Excel formula values), so two runs give byte-identical output; inputs are never
    written. Outputs are overwritten (required for idempotency).

Merge-back correctness note: composite is recomputed from Items rather than read from
the Composite tab because openpyxl cannot evaluate Excel formulas — a workbook never
re-opened in Excel would have no cached values. Recomputing is both robust and the thing
the acceptance check hand-verifies against the Items tab.
"""

import argparse
import csv
import datetime as _dt
import glob
import os
import re
import sys

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(HERE)  # ba-agent/
DATA_DIR = os.path.join(HERE, "data")

# R2 tags counted into the composite (must match the literals script 2's Composite tab
# COUNTIFS on Items!$H and script 3's harvest use).
R2_EXTENSION = "extension"
R2_NOVEL = "novel"

# Columns read (by NAME, never by letter) from the scoring workbook's Items sheet.
ITEMS_COLS = ["session_id", "participant_id", "item_index", "item_type", "apparatus_swap",
              "R1_kb_id", "R2_tag"]
# Columns read from the returned raters registry.
RATER_COLS = ["novel_id", "rate1", "rate2", "final_decision"]

# Scoring columns written to the FRONT of master_scored.csv (Qualtrics columns follow;
# any Qualtrics column whose name collides with one of these is prefixed 'q_').
SCORING_COLS = ["session_id", "participant_id", "agent_type", "matched_on",
                "R1_hits", "R2_extension", "R2_novel", "composite_total",
                "n_novel_ideas", "n_creative_confirmed"]


# ── value coercion ───────────────────────────────────────────────────────────
_TRUE_STRS = {"true", "1", "yes", "y", "t"}
_FALSE_STRS = {"false", "0", "no", "n", "f"}


def _blank(v) -> bool:
    return v is None or str(v).strip() == ""


def _as_bool(v):
    """Coerce an Excel cell to True/False/None. Accepts native bools and the common
    text spellings raters may type ('TRUE'/'FALSE'/'yes'/'no'/'1'/'0')."""
    if isinstance(v, bool):
        return v
    if _blank(v):
        return None
    s = str(v).strip().lower()
    if s in _TRUE_STRS:
        return True
    if s in _FALSE_STRS:
        return False
    return None


# ── key normalisation (trim, case, R_ prefix) ────────────────────────────────
def _norm(s) -> str:
    """Lowercase + strip ALL surrounding/internal whitespace. Empty for blanks."""
    return re.sub(r"\s+", "", str(s or "")).lower()


def _key_variants(s):
    """Candidate join keys for one id: the normalised form and, for Qualtrics-style
    'R_...' ids, the form with the leading 'r_' stripped — so a pasted id that lost or
    gained the R_ prefix still matches. Order = match preference."""
    n = _norm(s)
    if not n:
        return []
    out = [n]
    if n.startswith("r_"):
        out.append(n[2:])
    return out


# ── reading the scoring workbook (Items tab) ─────────────────────────────────
def read_scored_sessions(scoring_path):
    """Recompute per-session scores from the Items tab.

    NOTE this intentionally diverges from the Excel Composite tab's raw COUNTIFS
    formulas for rows where a rater filled BOTH R1_kb_id and R2_tag: the sheet's
    independent COUNTIFS would double-count such a row (once as an R1 hit, once as
    R2). Python instead enforces the documented rule that R1 and R2 are disjoint —
    an R1 hit is not also scored as R2 — and reports every such row via
    `both_filled` so it can be caught and fixed at the source (R2_tag is only
    supposed to be used on R1-blank rows; the dropdown doesn't hard-enforce that).

    Returns (sessions, order, both_filled):
      sessions    : dict session_id -> {participant_id, R1_hits, R2_extension,
                                        R2_novel, composite_total}
      order       : list of session_ids in first-seen row order (== Composite tab order)
      both_filled : list of (session_id, item_index) rows with both R1_kb_id and a
                    valid R2_tag filled — scored as R1 only; flag for the human.
    """
    wb = load_workbook(scoring_path, read_only=True, data_only=False)
    if "Items" not in wb.sheetnames:
        sys.exit(f"ERROR: {scoring_path} has no 'Items' sheet — is this a scoring "
                 f"workbook produced by 2_build_workbook.py?")
    ws = wb["Items"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        sys.exit(f"ERROR: {scoring_path} 'Items' sheet is empty.")
    col = {name: i for i, name in enumerate(header)}
    missing = [c for c in ITEMS_COLS if c not in col]
    if missing:
        sys.exit(f"ERROR: Items sheet missing column(s) {missing}. Found: {header}")

    sessions = {}
    order = []
    both_filled = []
    for r in rows:
        sid = r[col["session_id"]]
        if _blank(sid):
            continue  # every real item row carries a session_id
        sid = str(sid)
        if sid not in sessions:
            order.append(sid)
            sessions[sid] = {
                "participant_id": (r[col["participant_id"]] or "") if not _blank(
                    r[col["participant_id"]]) else "",
                "R1_hits": 0, "R2_extension": 0, "R2_novel": 0, "composite_total": 0,
            }
        rec = sessions[sid]
        # apparatus_swap rows are quarantined in BOTH rounds — never scored.
        if _as_bool(r[col["apparatus_swap"]]):
            continue
        tag = ("" if _blank(r[col["R2_tag"]]) else str(r[col["R2_tag"]]).strip().lower())
        if not _blank(r[col["R1_kb_id"]]):
            rec["R1_hits"] += 1
            if tag in (R2_EXTENSION, R2_NOVEL):
                both_filled.append((sid, r[col["item_index"]]))
            continue  # an R1 hit is not also an R2 tag (Composite counts them disjoint)
        if tag == R2_EXTENSION:
            rec["R2_extension"] += 1
        elif tag == R2_NOVEL:
            rec["R2_novel"] += 1
    wb.close()
    for rec in sessions.values():
        rec["composite_total"] = rec["R1_hits"] + rec["R2_extension"] + rec["R2_novel"]
    return sessions, order, both_filled


# ── reading the returned raters registry ─────────────────────────────────────
def read_final_decisions(raters_path):
    """Read the returned novel_ideas_raters workbook.

    Returns (decisions, rate_pairs, blanks):
      decisions  : dict novel_id -> bool (final_decision)
      rate_pairs : list of (rate1_bool, rate2_bool) over fully-rated rows (for kappa)
      blanks     : list of (novel_id, [which fields are blank]) — the validation gate
    """
    wb = load_workbook(raters_path, read_only=True, data_only=True)
    # script 3 names the sheet 'Ratings'; fall back to the first sheet defensively.
    ws = wb["Ratings"] if "Ratings" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        sys.exit(f"ERROR: {raters_path} rating sheet is empty.")
    col = {name: i for i, name in enumerate(header)}
    missing = [c for c in RATER_COLS if c not in col]
    if missing:
        sys.exit(f"ERROR: raters registry missing column(s) {missing}. Found: {header}")

    decisions = {}
    rate_pairs = []
    blanks = []
    for r in rows:
        nid = r[col["novel_id"]]
        if _blank(nid):
            continue
        nid = str(nid).strip()
        r1, r2, fd = r[col["rate1"]], r[col["rate2"]], r[col["final_decision"]]
        bad = []
        if _as_bool(r1) is None:
            bad.append("rate1")
        if _as_bool(r2) is None:
            bad.append("rate2")
        if _as_bool(fd) is None:
            bad.append("final_decision")
        if bad:
            blanks.append((nid, bad))
            continue
        decisions[nid] = _as_bool(fd)
        rate_pairs.append((_as_bool(r1), _as_bool(r2)))
    wb.close()
    return decisions, rate_pairs, blanks


# ── novel crosswalk (novel_id -> sessions) ───────────────────────────────────
def read_novel_crosswalk(path):
    """Return session_id -> set(novel_id) from script 3's novel_crosswalk.csv."""
    by_session = {}
    all_ids = set()
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd = csv.DictReader(f)
        need = {"novel_id", "session_id"}
        if not need.issubset(set(rd.fieldnames or [])):
            sys.exit(f"ERROR: {path} missing columns {need - set(rd.fieldnames or [])}. "
                     f"Found: {rd.fieldnames}")
        for row in rd:
            sid = (row.get("session_id") or "").strip()
            nid = (row.get("novel_id") or "").strip()
            if not sid or not nid:
                continue
            by_session.setdefault(sid, set()).add(nid)
            all_ids.add(nid)
    return by_session, all_ids


# ── crosswalk (session_id -> agent_type) ─────────────────────────────────────
def read_crosswalk(path):
    out = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd = csv.DictReader(f)
        if "session_id" not in (rd.fieldnames or []):
            sys.exit(f"ERROR: {path} has no 'session_id' column. Found: {rd.fieldnames}")
        for row in rd:
            sid = (row.get("session_id") or "").strip()
            if sid:
                out[sid] = (row.get("agent_type") or "").strip()
    return out


# ── Qualtrics export (CSV or XLSX) ───────────────────────────────────────────
def read_qualtrics(path, skip_after_header):
    """Return (header, rows) where rows is a list of dicts keyed by header.

    First row is the header. Qualtrics CSV exports insert extra descriptor rows
    (question text + an ImportId JSON) immediately after the header — pass
    --qualtrics-skip-rows 2 to drop them so they don't become phantom participants.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header = ["" if h is None else str(h) for h in next(it)]
        except StopIteration:
            sys.exit(f"ERROR: {path} is empty.")
        raw = [list(r) for r in it]
        wb.close()
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = list(csv.reader(f))
        if not reader:
            sys.exit(f"ERROR: {path} is empty.")
        header = reader[0]
        raw = reader[1:]

    for _ in range(max(0, skip_after_header)):
        if raw:
            raw.pop(0)

    rows = []
    for values in raw:
        rows.append({header[i]: (values[i] if i < len(values) else "")
                     for i in range(len(header))})
    return header, rows


# ── Cohen's kappa (binary, stdlib arithmetic) ────────────────────────────────
def cohens_kappa(pairs):
    """Return (raw_agreement, kappa) for a list of (bool, bool) rating pairs.
    kappa is None when there is no variance (pe == 1, kappa undefined)."""
    n = len(pairs)
    if n == 0:
        return None, None
    agree = sum(1 for a, b in pairs if a == b)
    po = agree / n
    p1 = sum(1 for a, _ in pairs if a) / n
    p2 = sum(1 for _, b in pairs if b) / n
    pe = p1 * p2 + (1 - p1) * (1 - p2)
    if pe >= 1.0:
        return po, None
    return po, (po - pe) / (1 - pe)


# ── file discovery ───────────────────────────────────────────────────────────
def _collection_from_scoring(path):
    m = re.match(r"scoring_(?P<coll>.+)_\d{4}-\d{2}-\d{2}\.xlsx$",
                 os.path.basename(path))
    return m.group("coll") if m else "unknown"


def _latest(pattern):
    cands = sorted(glob.glob(os.path.join(DATA_DIR, pattern)))
    return cands[-1] if cands else None


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--scoring", default=None,
                    help="Scored scoring_*.xlsx (default: newest in analysis/data).")
    ap.add_argument("--crosswalk", default=None,
                    help="crosswalk_*.csv (default: newest for the scoring collection).")
    ap.add_argument("--raters", default=None,
                    help="Returned novel_ideas_raters_*.xlsx (default: newest for "
                         "the collection). Omit only if there was no creativity stage.")
    ap.add_argument("--novel-crosswalk", default=None,
                    help="novel_crosswalk_*.csv (default: newest for the collection).")
    ap.add_argument("--snapshot", default=None,
                    help="raw_snapshot_*.json (default: newest for the collection) — "
                         "used only for the N(snapshot) reconciliation count.")
    ap.add_argument("--qualtrics", required=True,
                    help="Qualtrics export (CSV or XLSX). Never modified.")
    ap.add_argument("--responseid-col", default="ResponseId",
                    help="Qualtrics column holding ResponseId (default: ResponseId).")
    ap.add_argument("--session-col", default=None,
                    help="Qualtrics column(s) holding the participant-pasted Session ID; "
                         "enables the fallback join (omit to disable). Comma-separated "
                         "when split across per-arm fields, e.g. 'Q23,Q27,Q28' — the "
                         "first non-blank value in each row is used.")
    ap.add_argument("--qualtrics-skip-rows", type=int, default=0,
                    help="Descriptor rows to drop AFTER the header (Qualtrics CSV "
                         "usually needs 2: question text + ImportId JSON).")
    ap.add_argument("--out-dir", default=DATA_DIR, help="Output dir (default: analysis/data).")
    args = ap.parse_args()

    scoring_path = args.scoring or _latest("scoring_*.xlsx")
    if not scoring_path or not os.path.exists(scoring_path):
        sys.exit("ERROR: no scoring workbook found. Run 2_build_workbook.py first, or "
                 "pass --scoring <path>.")
    collection = _collection_from_scoring(scoring_path)

    crosswalk_path = args.crosswalk or _latest(f"crosswalk_{collection}_*.csv")
    if not crosswalk_path or not os.path.exists(crosswalk_path):
        sys.exit(f"ERROR: no crosswalk found for collection '{collection}'. Pass "
                 f"--crosswalk <path>.")

    raters_path = args.raters or _latest(f"novel_ideas_raters_{collection}_*.xlsx")
    novel_cross_path = args.novel_crosswalk or _latest(f"novel_crosswalk_{collection}_*.csv")
    snapshot_path = args.snapshot or _latest(f"raw_snapshot_{collection}_*.json")

    if not os.path.exists(args.qualtrics):
        sys.exit(f"ERROR: Qualtrics export not found: {args.qualtrics}")

    today = _dt.date.today().isoformat()
    out_csv = os.path.join(args.out_dir, f"master_scored_{collection}_{today}.csv")
    out_txt = os.path.join(args.out_dir, f"reconciliation_report_{collection}_{today}.txt")

    # ── load scores ──────────────────────────────────────────────────────────
    sessions, order, both_filled = read_scored_sessions(scoring_path)
    agent_of = read_crosswalk(crosswalk_path)

    # ── creativity stage (optional) ──────────────────────────────────────────
    creativity_on = bool(raters_path and os.path.exists(raters_path))
    novel_by_session, novel_ids_in_crosswalk = ({}, set())
    decisions, rate_pairs = {}, []
    creativity_notes = []
    if creativity_on:
        if not (novel_cross_path and os.path.exists(novel_cross_path)):
            sys.exit(f"ERROR: raters file {raters_path} was supplied but its "
                     f"novel_crosswalk_{collection}_*.csv is missing — the merge-back "
                     f"key store. Pass --novel-crosswalk <path>.")
        decisions, rate_pairs, blanks = read_final_decisions(raters_path)
        if blanks:
            lines = "\n".join(f"   {nid}: blank {', '.join(fields)}"
                              for nid, fields in blanks)
            sys.exit("ERROR: the returned raters registry has unfinished rows. Every "
                     "novel_id needs rate1, rate2 AND final_decision filled before "
                     f"merging.\n{lines}")
        novel_by_session, novel_ids_in_crosswalk = read_novel_crosswalk(novel_cross_path)
        # integrity: every novel_id in the crosswalk must have a decision and vice versa
        undecided = sorted(novel_ids_in_crosswalk - set(decisions))
        orphan = sorted(set(decisions) - novel_ids_in_crosswalk)
        if undecided:
            sys.exit("ERROR: novel_id(s) present in the crosswalk but absent from the "
                     f"returned raters file (cannot map creativity): {undecided}")
        if orphan:
            creativity_notes.append(
                f"raters file has {len(orphan)} novel_id(s) not in the crosswalk "
                f"(ignored): {orphan}")
    else:
        creativity_notes.append(
            "NO creativity stage inputs found (no raters file). n_novel_ideas and "
            "n_creative_confirmed are 0 for all rows; inter-rater agreement not reported.")

    # ── Qualtrics + join indexes ─────────────────────────────────────────────
    q_header, q_rows = read_qualtrics(args.qualtrics, args.qualtrics_skip_rows)
    if args.responseid_col not in q_header:
        sys.exit(f"ERROR: ResponseId column '{args.responseid_col}' not in Qualtrics "
                 f"header. Columns: {q_header}")
    # Session ID may be split across per-arm columns (e.g. Q23/Q27/Q28); only one is
    # filled per respondent, so we index every value and coalesce first-non-blank per row.
    session_cols = [c.strip() for c in (args.session_col or "").split(",") if c.strip()]
    bad_cols = [c for c in session_cols if c not in q_header]
    if bad_cols:
        sys.exit(f"ERROR: session column(s) {bad_cols} not in Qualtrics header. "
                 f"Columns: {q_header}")

    # index Qualtrics rows by normalised ResponseId and (optionally) pasted session id
    by_responseid = {}
    by_pasted_sid = {}
    for i, qr in enumerate(q_rows):
        for k in _key_variants(qr.get(args.responseid_col)):
            by_responseid.setdefault(k, i)
        for c in session_cols:
            for k in _key_variants(qr.get(c)):
                by_pasted_sid.setdefault(k, i)

    # Qualtrics column output names (prefix any that clash with scoring columns)
    q_out_name = {c: (f"q_{c}" if c in SCORING_COLS else c) for c in q_header}

    # ── join each scored session ─────────────────────────────────────────────
    matched_q_rows = set()
    master = []
    matched_primary = matched_fallback = unmatched = 0
    for sid in order:
        rec = sessions[sid]
        pid = rec["participant_id"]
        matched_on = "unmatched"
        q_idx = None
        for k in _key_variants(pid):
            if k in by_responseid:
                q_idx = by_responseid[k]
                matched_on = "participant_id"
                break
        if q_idx is None and session_cols:
            for k in _key_variants(sid):
                if k in by_pasted_sid:
                    q_idx = by_pasted_sid[k]
                    matched_on = "session_id"
                    break

        if matched_on == "participant_id":
            matched_primary += 1
        elif matched_on == "session_id":
            matched_fallback += 1
        else:
            unmatched += 1
        if q_idx is not None:
            matched_q_rows.add(q_idx)

        nids = novel_by_session.get(sid, set())
        n_novel = len(nids)
        n_conf = sum(1 for nid in nids if decisions.get(nid) is True)

        out = {
            "session_id": sid,
            "participant_id": pid,
            "agent_type": agent_of.get(sid, ""),
            "matched_on": matched_on,
            "R1_hits": rec["R1_hits"],
            "R2_extension": rec["R2_extension"],
            "R2_novel": rec["R2_novel"],
            "composite_total": rec["composite_total"],
            "n_novel_ideas": n_novel,
            "n_creative_confirmed": n_conf,
        }
        if q_idx is not None:
            for c in q_header:
                out[q_out_name[c]] = q_rows[q_idx].get(c, "")
        else:
            for c in q_header:
                out[q_out_name[c]] = ""
        master.append(out)

    # ── write master_scored.csv ──────────────────────────────────────────────
    fieldnames = SCORING_COLS + [q_out_name[c] for c in q_header]
    os.makedirs(args.out_dir, exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(master)

    # ── reconciliation report ────────────────────────────────────────────────
    n_snapshot = None
    if snapshot_path and os.path.exists(snapshot_path):
        import json
        with open(snapshot_path, encoding="utf-8") as f:
            n_snapshot = json.load(f).get("_meta", {}).get("n_docs")

    # Qualtrics rows that no scored session matched
    q_only = [i for i in range(len(q_rows)) if i not in matched_q_rows]
    scored_only = [m for m in master if m["matched_on"] == "unmatched"]
    fallback_rows = [m for m in master if m["matched_on"] == "session_id"]

    po = kappa = None
    if creativity_on:
        po, kappa = cohens_kappa(rate_pairs)

    L = []
    L.append("=" * 70)
    L.append("RECONCILIATION REPORT — 4_merge.py")
    L.append("=" * 70)
    L.append(f"generated        : {_dt.datetime.now().isoformat(timespec='seconds')}")
    L.append(f"collection       : {collection}")
    L.append(f"scoring workbook : {scoring_path}")
    L.append(f"crosswalk        : {crosswalk_path}")
    L.append(f"raters file      : {raters_path if creativity_on else '(none)'}")
    L.append(f"novel crosswalk  : {novel_cross_path if creativity_on else '(none)'}")
    L.append(f"qualtrics export : {args.qualtrics}")
    L.append(f"  responseid col : {args.responseid_col}")
    L.append(f"  session col(s) : {', '.join(session_cols) or '(none — fallback join disabled)'}")
    L.append("")
    L.append("COUNTS")
    L.append(f"  N(snapshot docs)      : {n_snapshot if n_snapshot is not None else 'n/a'}")
    L.append(f"  N(scored sessions)    : {len(order)}   (== master_scored rows)")
    L.append(f"  N(Qualtrics rows)     : {len(q_rows)}")
    L.append("")
    L.append("JOIN")
    L.append(f"  matched on participant_id (primary)  : {matched_primary}")
    L.append(f"  matched on session_id     (fallback) : {matched_fallback}")
    L.append(f"  unmatched                            : {unmatched}")
    if fallback_rows:
        L.append("  rows rescued by the session_id fallback:")
        for m in fallback_rows:
            L.append(f"     session={m['session_id']}  pid={m['participant_id'] or '(none)'}")
    if scored_only:
        L.append(f"  scored sessions with NO Qualtrics match ({len(scored_only)}) "
                 f"— explain (pilots/dropouts/blank pid):")
        for m in scored_only:
            L.append(f"     session={m['session_id']}  pid={m['participant_id'] or '(none)'}")
    if q_only:
        L.append(f"  Qualtrics rows with NO scored session ({len(q_only)}) "
                 f"— explain (no framework submitted/excluded):")
        for i in q_only:
            rid = q_rows[i].get(args.responseid_col, "")
            L.append(f"     ResponseId={rid}")
    L.append("")
    if both_filled:
        L.append(f"WARNING: {len(both_filled)} item row(s) have BOTH R1_kb_id AND a valid "
                 f"R2_tag filled. R2_tag is only meant to be used on R1-blank rows (the "
                 f"dropdown does not hard-enforce this). These rows were scored as R1 ONLY "
                 f"(R2 ignored) — this diverges from the Excel Composite tab's raw COUNTIFS, "
                 f"which would double-count them. Fix in scoring.xlsx (clear the stray "
                 f"R2_tag) and re-run, or confirm the R1_kb_id was a mistake:")
        for sid, idx in both_filled:
            L.append(f"     session={sid}  item_index={idx}")
        L.append("")
    L.append("CREATIVITY (standalone — does NOT affect composite)")
    if creativity_on:
        total_novel = sum(m["n_novel_ideas"] for m in master)
        total_conf = sum(m["n_creative_confirmed"] for m in master)
        L.append(f"  unique novel ideas rated       : {len(decisions)}")
        L.append(f"  final_decision == TRUE         : {sum(1 for v in decisions.values() if v)}")
        L.append(f"  sum n_novel_ideas (per-person) : {total_novel}")
        L.append(f"  sum n_creative_confirmed       : {total_conf}")
        L.append(f"  inter-rater raw agreement      : "
                 f"{po * 100:.1f}%  (n={len(rate_pairs)} ideas)")
        L.append(f"  Cohen's kappa                  : "
                 f"{kappa:.3f}" if kappa is not None
                 else "  Cohen's kappa                  : undefined (no rating variance)")
        L.append("  NOTE: raters worked SEQUENTIALLY on one shared file — rater 2 was "
                 "not blind to rater 1. Agreement is non-independent; report as such.")
    for note in creativity_notes:
        L.append(f"  {note}")
    L.append("")
    L.append("OUTPUTS")
    L.append(f"  master_scored : {out_csv}")
    L.append(f"  this report   : {out_txt}")
    L.append("  Inputs were not modified. Re-running regenerates identical output.")
    L.append("=" * 70)

    report = "\n".join(L)
    with open(out_txt, "w", encoding="utf-8") as f:
        f.write(report + "\n")
    print(report)


if __name__ == "__main__":
    main()
