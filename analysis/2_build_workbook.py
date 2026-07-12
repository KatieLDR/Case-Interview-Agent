#!/usr/bin/env python3
"""2_build_workbook.py — snapshot -> split into items -> blind scoring workbook.

Step 2 of the analysis-scoring pipeline (see ba-agent/docs/analysis_scoring_plan.md).

Reads an immutable snapshot produced by 1_pull_snapshot.py and emits, into
analysis/data/:
  - scoring_<collection>_<date>.xlsx   three tabs (KB_Reference / Items / Composite),
                                       with data-validation dropdowns. This is the file
                                       the human scores by hand, BLIND (no arm shown).
  - crosswalk_<collection>_<date>.csv  session_id, participant_id, agent_type. The ONLY
                                       place arm identity lives; used by 4_merge.py only.

Design guarantees (mirror the plan's acceptance checklist):
  * READ-ONLY on Firestore and on backend/ — this script never queries Firestore at all;
    it reads the frozen snapshot JSON. No .set()/.update()/.delete() anywhere.
  * NO LLM. Scoring is human judgement; this script only automates mechanics (splitting,
    KB seeding, dropdowns, apparatus_swap flagging). No genai/LLM imports.
  * BLIND. agent_type appears ONLY in crosswalk.csv, never in scoring.xlsx.
  * NON-DESTRUCTIVE. Refuses to overwrite an existing scoring.xlsx (it may already hold
    hand-entered scores) or crosswalk. Re-run with a fresh date stamp or remove the file.

The only backend import is backend.knowledge.knowledge_base (pure JSON read, no side
effects, no LLM) to seed the 35-row KB reference and read the swap concept's sub-bullets.
SWAP_CONFIG match_terms are MIRRORED below (not imported): backend/tools/concept_swap.py
imports backend.llm at module load, which is exactly the LLM dependency we must not pull.
"""

import argparse
import csv
import datetime as _dt
import glob
import json
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(HERE)  # ba-agent/
DATA_DIR = os.path.join(HERE, "data")

# Import the pure-JSON KB accessor (no LLM, no side effects). See module docstring.
sys.path.insert(0, REPO_ROOT)
from backend.knowledge import knowledge_base as kb  # noqa: E402


# ── source-ref cleaner ──────────────────────────────────────────────────────
# Mirror of backend/domain/grounding.py:_strip_source_refs — removes inline
# citation markers like " [a]" from item text. Replicated (not imported) to keep
# the analysis scripts' import surface minimal and obviously LLM-free.
_REF_RE = re.compile(r"\s*\[[a-z]\]")


def _strip_source_refs(text: str) -> str:
    return _REF_RE.sub("", text or "").strip()


# ── apparatus_swap match set ────────────────────────────────────────────────
# The planted distractor ("swap") is scored in NEITHER round — it is quarantined.
# An item is tagged apparatus_swap if its own text, OR its parent heading, matches
# any term below. The current swap name + its sub-bullets come from the KB (always
# in sync). match_terms are MIRRORED from backend/tools/concept_swap.py SWAP_CONFIG
# (identical across all three arms). PILOT_SWAP_TERMS cover the older g0 pilot
# wordings ("steps walked per day", etc.) that predate the current GenAI-use-cases
# swap; they only ever fire on the pilot snapshot and are harmless on g1.
#
# KEEP IN SYNC: if SWAP_CONFIG["*"]["match_terms"] changes in concept_swap.py,
# update SWAP_MATCH_TERMS below. (A backend gate already asserts match_terms track
# the KB swap entry; this mirror is the analysis-side copy.)
SWAP_MATCH_TERMS = [
    "use cases submitted", "use case submitted", "cases submitted",
    "submitted company-wide", "submitted company wide",
    "company-wide submission", "companywide submission",
    "submission count", "submission volume", "number of submissions",
    "submissions this year", "total number of use cases",
    "total use cases submitted", "innovation pipeline",
    "pipeline volume", "pipeline metric", "portfolio volume",
    "aggregate count", "aggregate number", "company-wide total",
    "total genai use cases", "number of genai use cases",
    "total number of genai use cases", "genai use case count",
    "genai use cases total", "genai submissions",
    "use case submissions", "submissions metric", "submission rate",
    "company-wide use case count", "companywide use case count",
]

PILOT_SWAP_TERMS = [
    "steps walked per day", "steps walked", "number of steps walked",
    "fears of older colleagues", "older colleagues regarding replacement",
    "does the it team currently log this data",
    "baseline indicator for measuring the impact",
    "baseline indicator for genai rollout",
]


def _swap_terms():
    """Full lowercase term set: KB swap name + KB sub-bullets + match_terms + pilot."""
    terms = list(SWAP_MATCH_TERMS) + list(PILOT_SWAP_TERMS)
    swap = kb.get_swap_concept()
    if swap:
        if swap.get("name"):
            terms.append(swap["name"])
        for b in swap.get("sub_bullets", []):
            terms.append(_strip_source_refs(b))
    return [t.lower().strip() for t in terms if t and t.strip()]


def _is_swap_text(text: str, terms) -> bool:
    t = (text or "").lower()
    return any(term in t for term in terms)


# ── splitter ────────────────────────────────────────────────────────────────
_TITLE = "final framework summary"
_HEADING_RE = re.compile(r"^\*\*(?P<txt>.+?)\*\*$")
_BULLET_RE = re.compile(r"^-\s+(?P<txt>.+)$")
_APPROVED_RE = re.compile(r"^approved concepts\s*:\s*(?P<rest>.+)$", re.IGNORECASE)


def _is_italic_instruction(s: str) -> bool:
    """Standalone fully-italic line, e.g. '*Add or remove concepts as you like*'."""
    return (
        len(s) >= 2
        and s.startswith("*")
        and s.endswith("*")
        and not s.startswith("**")
        and not s.endswith("**")
    )


def split_items(text: str):
    """Split a framework artefact into ordered items.

    Locked rules (from the plan):
      **X**        -> heading item (skip the 'Final Framework Summary' title)
      - X          -> bullet item (attributed to the current heading)
      *...*        -> italic instruction line, skipped
      preamble     -> any non-heading/non-bullet prose (incl. before the first
                      heading) is skipped
      'Approved concepts: a, b, c' -> comma-split into concept items (flat fallback)

    Returns a list of {item_type, item_text, parent_heading}.
    """
    items = []
    current_heading = ""
    for raw in (text or "").split("\n"):
        s = raw.strip()
        if not s:
            continue

        m = _APPROVED_RE.match(s)
        if m:
            for part in m.group("rest").split(","):
                p = _strip_source_refs(part)
                if p:
                    items.append({"item_type": "concept", "item_text": p,
                                  "parent_heading": ""})
            continue

        if _is_italic_instruction(s):
            continue

        m = _HEADING_RE.match(s)
        if m:
            inner = _strip_source_refs(m.group("txt"))
            if not inner or inner.lower() == _TITLE:
                continue  # skip the summary title
            current_heading = inner
            items.append({"item_type": "heading", "item_text": inner,
                          "parent_heading": ""})
            continue

        m = _BULLET_RE.match(s)
        if m:
            btxt = _strip_source_refs(m.group("txt"))
            if btxt:
                items.append({"item_type": "bullet", "item_text": btxt,
                              "parent_heading": current_heading})
            continue

        # anything else (preamble, stray prose) -> skipped
    return items


# ── per-doc artefact source selection ───────────────────────────────────────
def select_source(doc):
    """Pick the artefact text per arm and note truncation loss.

    XAI : current_answer
    BB  : current_answer -> final_framework (truncated fallback)
    HITL: current_answer -> concepts_approved list

    Returns (source_name, items_list, truncation_loss_bool) or (None, [], False).
    """
    at = doc.get("agent_type")
    ca = doc.get("current_answer")
    if ca and ca.strip():
        items = split_items(ca)
        if items:
            return "current_answer", items, False
        # current_answer present but degenerate (e.g. only the title, no items) —
        # fall through to the arm's fallback source below.

    if at == "black_box":
        ff = doc.get("final_framework")
        if ff and ff.strip():
            # BB truncates final_framework to 1000 chars (black_box.py:551). A doc
            # falling back to a 1000-char final_framework with null current_answer
            # has lost content — flag it for manual review.
            loss = len(ff) >= 1000
            return "final_framework", split_items(ff), loss

    if at == "hitl":
        approved = doc.get("concepts_approved")
        if approved:
            items = [{"item_type": "concept", "item_text": _strip_source_refs(str(c)),
                      "parent_heading": ""}
                     for c in approved if str(c).strip()]
            return "concepts_approved", items, False

    return None, [], False


# ── dedup ───────────────────────────────────────────────────────────────────
def _dedup_key(doc):
    """Prefer a doc that ended; then the latest created_at."""
    ended = doc.get("ended_at")
    return (ended not in (None, ""), ended or "", doc.get("created_at") or "")


def dedup_by_pid(usable):
    """Collapse docs sharing a non-null participant_id to one; keep None-pid docs
    each as their own row. Returns (kept_docs, duplicate_report, tie_report)."""
    by_pid = {}
    none_pid = []
    for doc in usable:
        pid = doc.get("participant_id")
        if pid in (None, ""):
            none_pid.append(doc)
        else:
            by_pid.setdefault(pid, []).append(doc)

    kept = list(none_pid)
    dup_report = []
    tie_report = []
    for pid, docs in by_pid.items():
        if len(docs) == 1:
            kept.append(docs[0])
            continue
        ranked = sorted(docs, key=_dedup_key, reverse=True)
        winner = ranked[0]
        kept.append(winner)
        dup_report.append((pid, len(docs), winner.get("doc_id")))
        # tie = two docs indistinguishable on (ended, ended_at, created_at)
        if _dedup_key(ranked[0]) == _dedup_key(ranked[1]):
            tie_report.append((pid, [d.get("doc_id") for d in ranked]))
    return kept, dup_report, tie_report


# ── KB reference rows (5 pillars + 30 concepts = 35; swap excluded) ─────────
def kb_rows():
    rows = []
    for p in kb.get_all_pillars():
        rows.append({
            "kb_id": p["id"],
            "number": "",
            "name": p["name"],
            "pillar": p["name"],
            "status": "shown" if p.get("shown") else "withheld",
        })
    concepts = [c for c in kb.get_all_concepts() if not c.get("swap")]
    concepts.sort(key=lambda c: c.get("number", 0))
    pillar_name = {p["id"]: p["name"] for p in kb.get_all_pillars()}
    for c in concepts:
        rows.append({
            "kb_id": c["id"],
            "number": c.get("number", ""),
            "name": c["name"],
            "pillar": pillar_name.get(c.get("pillar_id"), ""),
            "status": "shown" if c.get("shown") else "withheld",
        })
    return rows


# ── workbook construction ───────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_SWAP_FILL = PatternFill("solid", fgColor="FCE4D6")
_HEADER_FONT = Font(bold=True)

ITEMS_HEADERS = ["session_id", "participant_id", "item_index", "item_type",
                 "item_text", "apparatus_swap", "R1_kb_id", "R2_tag", "notes"]


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def build_workbook(item_rows, kbrows, out_path):
    wb = Workbook()

    # 1) KB_Reference -------------------------------------------------------
    ws_kb = wb.active
    ws_kb.title = "KB_Reference"
    kb_headers = ["kb_id", "number", "name", "pillar", "status"]
    ws_kb.append(kb_headers)
    for r in kbrows:
        ws_kb.append([r["kb_id"], r["number"], r["name"], r["pillar"], r["status"]])
    _style_header(ws_kb, len(kb_headers))
    for col, w in zip("ABCDE", (26, 8, 46, 22, 10)):
        ws_kb.column_dimensions[col].width = w
    kb_last = len(kbrows) + 1  # row index of last kb id

    # 2) Items --------------------------------------------------------------
    ws = wb.create_sheet("Items")
    ws.append(ITEMS_HEADERS)
    for row in item_rows:
        ws.append([
            row["session_id"], row["participant_id"], row["item_index"],
            row["item_type"], row["item_text"],
            True if row["apparatus_swap"] else False,
            "", "",  # R1_kb_id, R2_tag left blank for the human
            row["notes"],
        ])
    _style_header(ws, len(ITEMS_HEADERS))
    for col, w in zip("ABCDEFGHI", (26, 18, 9, 10, 70, 14, 22, 12, 34)):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["E"].width = 70
    for r in range(2, len(item_rows) + 2):
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
        # tint pre-flagged swap rows so the human sees the quarantine at a glance
        if item_rows[r - 2]["apparatus_swap"]:
            for col in range(1, len(ITEMS_HEADERS) + 1):
                ws.cell(row=r, column=col).fill = _SWAP_FILL

    last = len(item_rows) + 1
    if last >= 2:
        # R1_kb_id dropdown: the 35 KB ids (col A of KB_Reference) + blank.
        dv_kb = DataValidation(
            type="list",
            formula1=f"KB_Reference!$A$2:$A${kb_last}",
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )
        dv_kb.error = "Pick a kb_id from the KB_Reference tab, or leave blank."
        dv_kb.errorTitle = "Not a KB id"
        dv_kb.prompt = "Round 1: KB id this item matches. Blank => score in Round 2."
        dv_kb.promptTitle = "R1_kb_id"
        ws.add_data_validation(dv_kb)
        dv_kb.add(f"G2:G{last}")

        # R2_tag dropdown: only meaningful on R1-blank rows.
        dv_tag = DataValidation(
            type="list",
            formula1='"extension,novel,excluded"',
            allow_blank=True,
            showErrorMessage=True,
            showInputMessage=True,
        )
        dv_tag.error = "Choose extension, novel, or excluded."
        dv_tag.errorTitle = "Not a valid tag"
        dv_tag.prompt = "Round 2 (only when R1_kb_id is blank): extension / novel / excluded."
        dv_tag.promptTitle = "R2_tag"
        ws.add_data_validation(dv_tag)
        dv_tag.add(f"H2:H{last}")

    ws.cell(row=1, column=6).comment = Comment(
        "TRUE = planted swap distractor. Pre-flagged; do NOT score (neither round).",
        "pipeline")
    ws.cell(row=1, column=7).comment = Comment(
        "Round 1: choose the KB id this item matches. Leave blank if no match "
        "(then tag it in Round 2 / R2_tag).", "pipeline")
    ws.cell(row=1, column=8).comment = Comment(
        "Round 2: fill ONLY for R1-blank, non-swap rows. "
        "extension = builds on a KB concept; novel = genuinely new; excluded = off-topic.",
        "pipeline")

    # 3) Composite ----------------------------------------------------------
    # Keyed on session_id (unique per doc) so the 46 None-pid pilot rows don't
    # collapse together. Formulas recompute live as the human fills R1/R2.
    ws_c = wb.create_sheet("Composite")
    comp_headers = ["session_id", "participant_id", "R1_hits",
                    "R2_extension", "R2_novel", "composite_total"]
    ws_c.append(comp_headers)
    # one Composite row per distinct session_id in Items, in first-seen order
    seen = []
    seen_set = set()
    pid_of = {}
    for row in item_rows:
        sid = row["session_id"]
        if sid not in seen_set:
            seen_set.add(sid)
            seen.append(sid)
            pid_of[sid] = row["participant_id"]
    IB, IF, IG, IH = "Items!$B", "Items!$F", "Items!$G", "Items!$H"
    ISID = "Items!$A"
    for i, sid in enumerate(seen):
        r = i + 2
        a = f"$A{r}"  # this row's session_id cell
        # R1 hit  = non-blank R1_kb_id, not swap
        r1 = f'=COUNTIFS({ISID}:$A,{a},{IG}:$G,"<>",{IF}:$F,FALSE)'
        r2e = f'=COUNTIFS({ISID}:$A,{a},{IH}:$H,"extension",{IF}:$F,FALSE)'
        r2n = f'=COUNTIFS({ISID}:$A,{a},{IH}:$H,"novel",{IF}:$F,FALSE)'
        total = f"=C{r}+D{r}+E{r}"
        ws_c.append([sid, pid_of[sid], r1, r2e, r2n, total])
    _style_header(ws_c, len(comp_headers))
    for col, w in zip("ABCDEF", (26, 18, 10, 14, 10, 16)):
        ws_c.column_dimensions[col].width = w

    wb.save(out_path)
    return len(seen)


# ── main ────────────────────────────────────────────────────────────────────
def _latest_snapshot():
    cands = sorted(glob.glob(os.path.join(DATA_DIR, "raw_snapshot_*.json")))
    return cands[-1] if cands else None


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--snapshot", default=None,
                    help="Path to raw_snapshot_*.json (default: newest in analysis/data).")
    ap.add_argument("--out-dir", default=DATA_DIR,
                    help="Output directory (default: analysis/data).")
    args = ap.parse_args()

    snap_path = args.snapshot or _latest_snapshot()
    if not snap_path or not os.path.exists(snap_path):
        sys.exit("ERROR: no snapshot found. Run 1_pull_snapshot.py first, or pass "
                 "--snapshot <path>.")
    with open(snap_path, encoding="utf-8") as f:
        snap = json.load(f)
    meta = snap.get("_meta", {})
    collection = meta.get("collection", "unknown")
    docs = snap.get("docs", [])
    print(f"Snapshot   : {snap_path}")
    print(f"Collection : {collection}   (docs in file: {len(docs)})")

    # naming: encode collection + date so g0/g1 outputs can never be confused
    today = _dt.date.today().isoformat()
    xlsx_path = os.path.join(args.out_dir, f"scoring_{collection}_{today}.xlsx")
    cross_path = os.path.join(args.out_dir, f"crosswalk_{collection}_{today}.csv")
    for p in (xlsx_path, cross_path):
        if os.path.exists(p):
            sys.exit(f"ERROR: {p} already exists. The scoring workbook may hold "
                     f"hand-entered scores — it is never overwritten. Remove it "
                     f"manually to rebuild, or wait for tomorrow's date stamp.")

    terms = _swap_terms()

    # select source + split every doc that has a usable artefact
    usable = []
    trunc_docs = []
    for doc in docs:
        source, items, loss = select_source(doc)
        if source is None:
            continue
        doc["_source"] = source
        doc["_items"] = items
        if loss:
            trunc_docs.append(doc.get("doc_id"))
        usable.append(doc)

    kept, dup_report, tie_report = dedup_by_pid(usable)

    # build item rows (blind: no agent_type)
    item_rows = []
    swap_item_count = 0
    zero_item_docs = []
    for doc in kept:
        sid = doc.get("doc_id")
        pid = doc.get("participant_id") or ""
        if not doc["_items"]:
            # usable source but nothing survived the splitter (e.g. an empty
            # framework) — surfaced below, never dropped without a trace.
            zero_item_docs.append((sid, doc.get("agent_type"), doc.get("_source")))
        for idx, it in enumerate(doc["_items"], start=1):
            # Tag on the item's OWN text only. The canonical swap sub-bullets each
            # carry their own match anchor (KB sub_bullets + pilot terms), so we do
            # NOT propagate a swap heading's flag to its children — a participant can
            # file legitimate KB concepts under a swap-ish heading, and those must
            # stay scoreable rather than be quarantined by association.
            is_swap = _is_swap_text(it["item_text"], terms)
            if is_swap:
                swap_item_count += 1
            item_rows.append({
                "session_id": sid,
                "participant_id": pid,
                "item_index": idx,
                "item_type": it["item_type"],
                "item_text": it["item_text"],
                "apparatus_swap": is_swap,
                "notes": "apparatus_swap — do not score" if is_swap else "",
            })

    n_participants = build_workbook(item_rows, kb_rows(), xlsx_path)

    # crosswalk.csv — the ONLY file carrying agent_type
    with open(cross_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["session_id", "participant_id", "agent_type"])
        for doc in kept:
            w.writerow([doc.get("doc_id"), doc.get("participant_id") or "",
                        doc.get("agent_type") or ""])

    # ── data-quality report ────────────────────────────────────────────────
    none_pid = sum(1 for d in kept if (d.get("participant_id") in (None, "")))
    pid_samples = [d.get("participant_id") for d in kept
                   if d.get("participant_id") not in (None, "")][:5]
    # swap sanity: detected sessions should carry NO swap items; not-detected should
    detected_with_swap = []
    notdetected_no_swap = []
    swap_sids = {r["session_id"] for r in item_rows if r["apparatus_swap"]}
    for doc in kept:
        sid = doc.get("doc_id")
        if doc.get("concept_swap_detected") is True and sid in swap_sids:
            detected_with_swap.append(sid)
        if doc.get("concept_swap_detected") is False and sid not in swap_sids:
            notdetected_no_swap.append(sid)

    print("\n" + "=" * 70)
    print("DATA QUALITY")
    print("=" * 70)
    print(f"usable-source docs        : {len(usable)}")
    print(f"participants (post-dedup) : {len(kept)}  (Composite rows: {n_participants})")
    print(f"total items               : {len(item_rows)}")
    print(f"apparatus_swap items      : {swap_item_count} "
          f"(across {len(swap_sids)} sessions)")
    print(f"None/empty pid            : {none_pid}  (kept, keyed by session_id)")
    print(f"pid format sample         : {pid_samples}  (expect Qualtrics 'R_...')")
    if trunc_docs:
        print(f"BB truncation-loss docs   : {len(trunc_docs)} -> {trunc_docs}  "
              f"(1000-char final_framework fallback; review manually)")
    else:
        print("BB truncation-loss docs   : 0")
    if zero_item_docs:
        print(f"zero-item docs            : {len(zero_item_docs)} "
              f"(usable source, but splitter produced no items — NOT scored, listed):")
        for sid, arm, src in zero_item_docs:
            print(f"   {sid}  (src={src})")
    else:
        print("zero-item docs            : 0")
    if dup_report:
        print(f"\nduplicate pids resolved   : {len(dup_report)} "
              f"(kept newest ended/created; NONE dropped silently)")
        for pid, n, winner in dup_report:
            print(f"   {pid}: {n} docs -> kept {winner}")
    else:
        print("\nduplicate pids            : none")
    if tie_report:
        print("\n*** TIES needing manual resolution (indistinguishable timestamps):")
        for pid, ids in tie_report:
            print(f"   {pid}: {ids}")
    # swap tagger sanity (verification step 2 of the plan)
    print(f"\nswap sanity: {len(notdetected_no_swap)} not-detected sessions carry NO "
          f"swap item (some not-detected sessions should carry swap lines)")
    if detected_with_swap:
        print(f"swap sanity: {len(detected_with_swap)} session(s) marked "
              f"concept_swap_detected=TRUE still contain a swap-tagged item — INSPECT "
              f"(participant may have re-added swap phrasing after detecting): "
              f"{detected_with_swap[:10]}")
    else:
        print("swap sanity: no detected-session retains a swap item (good)")

    print("\n" + "=" * 70)
    print(f"Wrote workbook : {xlsx_path}")
    print(f"Wrote crosswalk: {cross_path}")
    print("scoring.xlsx is BLIND (no agent_type). agent_type lives only in crosswalk.")
    print("=" * 70)


if __name__ == "__main__":
    main()
